"""Extract full FSIA File Mapping from the ICD-D12 Excel spec."""
import openpyxl
import os

excel_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "doc", "specs", "WI DHS MES CMM_ICD-D12 FSIA File_v2.0_Unsubmitted Updates.xlsx"
)

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Extract the FSIA File Mapping sheet - just key columns
ws = wb["FSIA File Mapping"]
print(f"Total rows: {ws.max_row}")
print()

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    label = str(row[0])[:40] if row[0] else ""
    data_element = str(row[4])[:40] if row[4] else ""
    data_type = str(row[5])[:10] if row[5] else ""
    length = str(row[6])[:5] if row[6] else ""
    required = str(row[7])[:3] if row[7] else ""

    # Skip delimiter rows and empty rows
    if not label or label == "Delimiter":
        continue

    print(f"{row_idx:3d} | {label:40s} | {data_element:40s} | {data_type:6s} | {length:5s} | {required}")
