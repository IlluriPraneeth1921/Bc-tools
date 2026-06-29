"""Extract content from the ICD-D12 Excel specification."""
import openpyxl
import os

excel_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "doc", "specs", "WI DHS MES CMM_ICD-D12 FSIA File_v2.0_Unsubmitted Updates.xlsx"
)

wb = openpyxl.load_workbook(excel_path, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*80}")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True), 1):
        cells = [str(c)[:60] if c else "" for c in row]
        # Skip fully empty rows
        if any(cells):
            print(f"  Row {row_idx}: {cells}")
